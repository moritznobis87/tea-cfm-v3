"""
Tests des Pipeline-Ergebnis-Exports (engine/io_ergebnis_excel.py):
Blattstruktur, native Diagramme je Reiter, Uebersichtswerte,
Kollisionsbehandlung der Blattnamen und Markierung inaktiver Projekte.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from engine import pipeline_ergebnis_excel
from tests.conftest import _baue_global_assumptions, _baue_projekt

PROJEKT_NAME = _baue_projekt().name


@pytest.fixture(scope="module")
def arbeitsmappe():
    projekt = _baue_projekt()
    inaktiv = _baue_projekt()
    inaktiv.id = "p2"
    inaktiv.aktiv = False
    daten = pipeline_ergebnis_excel(
        [(projekt, projekt.name), (inaktiv, inaktiv.name)],
        _baue_global_assumptions(), n_mc=40,
    )
    return load_workbook(io.BytesIO(daten))


class TestPipelineErgebnisExport:
    def test_blaetter_und_namenskollision(self, arbeitsmappe):
        namen = arbeitsmappe.sheetnames
        assert namen[0] == "Übersicht"
        assert len(namen) == 3
        # Gleicher Wunschname -> eindeutige Blattnamen
        assert len(set(namen)) == 3

    def test_native_diagramme_je_projektblatt(self, arbeitsmappe):
        for name in arbeitsmappe.sheetnames[1:]:
            ws = arbeitsmappe[name]
            # Alle Auswertungen als native Charts (11 je Projekt)
            assert len(ws._charts) == 11
            # Diagramme referenzieren Zellbereiche (kein Bild)
            for chart in ws._charts:
                assert chart.series, name
        assert len(arbeitsmappe["Übersicht"]._charts) == 1

    def test_uebersicht_kennzeichnet_inaktive(self, arbeitsmappe):
        ws = arbeitsmappe["Übersicht"]
        werte = [
            [ws.cell(row=r, column=c).value for c in range(1, 10)]
            for r in range(3, 6)
        ]
        kopf = werte[0]
        assert "Aktiv" in kopf and "EK-Rendite" in kopf
        aktiv_spalte = kopf.index("Aktiv")
        assert {z[aktiv_spalte] for z in werte[1:]} == {"ja", "nein"}

    def test_kpi_und_tabellen_im_projektblatt(self, arbeitsmappe):
        ws = arbeitsmappe[arbeitsmappe.sheetnames[1]]
        assert PROJEKT_NAME in str(ws.cell(row=1, column=1).value)
        texte = {str(ws.cell(row=r, column=1).value) for r in range(1, 500)}
        for begriff in ("EK-Rendite (IRR)", "Erlösstruktur (€/Jahr)",
                        "Kapitaldienst und DSCR",
                        "Marktpreisszenarien im Vergleich"):
            assert any(begriff in t for t in texte), begriff
