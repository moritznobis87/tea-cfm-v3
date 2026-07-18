"""
Ergebnis-Export der gesamten Projekt-Pipeline als Excel-Arbeitsmappe:
ein Uebersichtsblatt plus je Projekt ein eigener Reiter mit allen
Dashboard-Auswertungen als NATIVE Excel-Diagramme (openpyxl-Charts,
keine Bilder) - jeweils mit dem Datengeruest als Tabelle daneben, so
dass die Diagramme in Excel voll editierbar bleiben.

Reine Werte-Ausgabe (Ergebnisbericht, keine Eingabezellen/Formeln);
Schrift Arial, deutsche Beschriftungen.
"""

from __future__ import annotations

import io

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

from .analytics import (
    MC_STANDARD_SIGMAS,
    run_monte_carlo,
    run_scenario_comparison,
    run_tornado,
)
from .models import AnlagenTyp, GlobalAssumptions, PVProject
from .pipeline import run_valuation
from .sensitivity import run_eag_sensitivity

try:  # Sprachtexte: eigenstaendiges Modul ohne Engine-Abhaengigkeit
    from texte import excel_texte
except ImportError:  # pragma: no cover - Fallback bei isoliertem Aufruf
    def excel_texte() -> dict[str, str]:
        return {}


def _t(schluessel: str, **platzhalter) -> str:
    """Liest einen Excel-Beschriftungstext; faellt bei fehlendem
    Schluessel auf den Schluessel selbst zurueck (fail-soft)."""
    text = excel_texte().get(schluessel, schluessel)
    return text.format(**platzhalter) if platzhalter else text

_FONT = "Arial"
_INK = "143530"
_BRAND = "BE172B"
_WASH = "F1F3F2"

_F_TITEL = Font(name=_FONT, size=14, bold=True, color=_INK)
_F_H2 = Font(name=_FONT, size=11, bold=True, color=_INK)
_F_KOPF = Font(name=_FONT, size=9, bold=True, color="FFFFFF")
_F_TEXT = Font(name=_FONT, size=9)
_FILL_KOPF = PatternFill("solid", fgColor=_INK)
_FILL_KPI = PatternFill("solid", fgColor=_WASH)

_EUR = "#,##0"
_CT = "0.00"
_PCT = "0.0%"

#: Diagrammgroesse (Zellenraster) und Spalte, in der Diagramme haengen.
_CHART_BREITE = 15.5
_CHART_HOEHE = 8.2
_CHART_SPALTE = "J"


def _stil(chart, titel: str, y_titel: str = "") -> None:
    chart.title = titel
    chart.style = 10
    chart.height = _CHART_HOEHE
    chart.width = _CHART_BREITE
    if y_titel:
        chart.y_axis.title = y_titel
    chart.y_axis.numFmt = "#,##0"


def _schreibe_tabelle(ws, zeile: int, df: pd.DataFrame,
                      formate: dict[str, str] | None = None) -> tuple[int, int]:
    """Schreibt einen DataFrame ab `zeile` (Spalte A) mit Kopfzeile;
    liefert (kopfzeile, letzte_zeile)."""
    formate = formate or {}
    for j, spalte in enumerate(df.columns, start=1):
        zelle = ws.cell(row=zeile, column=j, value=str(spalte))
        zelle.font = _F_KOPF
        zelle.fill = _FILL_KOPF
        zelle.alignment = Alignment(horizontal="center")
    for i, (_, r) in enumerate(df.iterrows(), start=1):
        for j, spalte in enumerate(df.columns, start=1):
            wert = r[spalte]
            if isinstance(wert, (np.floating, np.integer)):
                wert = wert.item()
            zelle = ws.cell(row=zeile + i, column=j, value=wert)
            zelle.font = _F_TEXT
            if spalte in formate:
                zelle.number_format = formate[spalte]
    return zeile, zeile + len(df)


def _sektion(ws, zeile: int, titel: str) -> int:
    ws.cell(row=zeile, column=1, value=titel).font = _F_H2
    return zeile + 1


def _haenge_chart(ws, chart, anker_zeile: int) -> None:
    ws.add_chart(chart, f"{_CHART_SPALTE}{anker_zeile}")


def _naechste_zeile(tab_ende: int, anker: int) -> int:
    """Naechster Sektionsstart: unter Tabelle UND Diagramm (~17 Zeilen)."""
    return max(tab_ende, anker + 17) + 2


def _linien_chart(ws, kopf: int, ende: int, spalten: list[int],
                  titel: str, y_titel: str) -> LineChart:
    chart = LineChart()
    _stil(chart, titel, y_titel)
    for c in spalten:
        chart.add_data(
            Reference(ws, min_col=c, min_row=kopf, max_row=ende),
            titles_from_data=True,
        )
    chart.set_categories(Reference(ws, min_col=1, min_row=kopf + 1, max_row=ende))
    return chart


def _balken_chart(ws, kopf: int, ende: int, spalten: list[int], titel: str,
                  y_titel: str, gestapelt: bool = False,
                  horizontal: bool = False) -> BarChart:
    chart = BarChart()
    chart.type = "bar" if horizontal else "col"
    if gestapelt:
        chart.grouping = "stacked"
        chart.overlap = 100
    _stil(chart, titel, y_titel)
    for c in spalten:
        chart.add_data(
            Reference(ws, min_col=c, min_row=kopf, max_row=ende),
            titles_from_data=True,
        )
    chart.set_categories(Reference(ws, min_col=1, min_row=kopf + 1, max_row=ende))
    return chart


def _blattname(name: str, vergeben: set[str]) -> str:
    for zeichen in "[]:*?/\\":
        name = name.replace(zeichen, " ")
    name = name.strip()[:28] or "Projekt"
    kandidat, i = name, 2
    while kandidat in vergeben:
        kandidat = f"{name[:25]} ({i})"
        i += 1
    vergeben.add(kandidat)
    return kandidat


# ---------------------------------------------------------------------------
# Projektblatt
# ---------------------------------------------------------------------------


def _projektblatt(ws, project: PVProject, ga: GlobalAssumptions,
                  n_mc: int) -> None:
    result = run_valuation(project, ga)
    df = result.cashflow.data
    kpis = result.kpis
    betrieb = df[df["jahr"] >= 1]

    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    for sp in "BCDEFGH":
        ws.column_dimensions[sp].width = 14

    typ = (_t("typ_agri") if project.anlagentyp == AnlagenTyp.AGRI_PV
           else _t("typ_konventionell"))
    titel = ws.cell(row=1, column=1, value=f"{project.name} ({typ})"
                    + ("" if project.aktiv else " – INAKTIV"))
    titel.font = _F_TITEL

    # --- KPI-Block ---
    kpi_paare = [
        (_t("kpi_nennleistung"), project.nennleistung_kwp, "#,##0 \"kWp\""),
        (_t("kpi_eag_zuschlagswert"), project.eag_zuschlagswert_ct_kwh, "0.00 \"ct/kWh\""),
        (_t("kpi_capex_gesamt"), kpis.capex_total_eur, _EUR + " €"),
        (_t("kpi_eigenkapital"), kpis.eigenkapital_eur, _EUR + " €"),
        (_t("kpi_ek_rendite"), kpis.equity_irr, _PCT),
        (_t("kpi_npv"), kpis.npv_eur, _EUR + " €"),
        (_t("kpi_dscr_min"), kpis.dscr_min, "0.00"),
        (_t("kpi_amortisation"), kpis.payback_jahre, "0.0 \"Jahre\""),
    ]
    for i, (name, wert, fmt) in enumerate(kpi_paare):
        z = 3 + i
        ws.cell(row=z, column=1, value=name).font = _F_TEXT
        ws.cell(row=z, column=1).fill = _FILL_KPI
        zelle = ws.cell(row=z, column=2, value=wert)
        zelle.font = Font(name=_FONT, size=9, bold=True)
        zelle.fill = _FILL_KPI
        zelle.number_format = fmt

    zeile = 13

    # --- 1. Verguetung & Marktwert ---
    zeile = _sektion(ws, zeile, _t("sektion_verguetung"))
    t = betrieb[["jahr", "marktwert_nominal_ct_kwh", "verguetungssatz_ct_kwh"]]
    t = t.rename(columns={"jahr": _t("spalte_jahr"),
                          "marktwert_nominal_ct_kwh": _t("spalte_marktwert_nominal"),
                          "verguetungssatz_ct_kwh": _t("spalte_verguetungssatz")})
    kopf, ende = _schreibe_tabelle(ws, zeile, t, {
        _t("spalte_marktwert_nominal"): _CT, _t("spalte_verguetungssatz"): _CT})
    chart = _linien_chart(ws, kopf, ende, [2, 3],
                          _t("chart_verguetung_titel"), "ct/kWh")
    chart.y_axis.numFmt = "0.00"
    _haenge_chart(ws, chart, kopf)
    zeile = _naechste_zeile(ende, kopf)

    # --- 2. Erloesstruktur ---
    zeile = _sektion(ws, zeile, _t("sektion_erloesstruktur"))
    t = betrieb[["jahr", "erloes_markt_eur", "erloes_praemie_eur"]].rename(
        columns={"jahr": _t("spalte_jahr"), "erloes_markt_eur": _t("spalte_markterloes"),
                 "erloes_praemie_eur": _t("spalte_marktpraemie")})
    kopf, ende = _schreibe_tabelle(ws, zeile, t,
                                   {_t("spalte_markterloes"): _EUR, _t("spalte_marktpraemie"): _EUR})
    _haenge_chart(ws, _balken_chart(ws, kopf, ende, [2, 3],
                                    _t("chart_erloesstruktur_titel"), "€/Jahr", gestapelt=True),
                  kopf)
    zeile = _naechste_zeile(ende, kopf)

    # --- 3. Gesamt-Cashflow (Kombi Balken + kumulierte Linie) ---
    zeile = _sektion(ws, zeile, _t("sektion_cashflow"))
    t = df[["jahr", "cf_gesamt_eur", "cf_kumuliert_eur"]].rename(
        columns={"jahr": _t("spalte_jahr"), "cf_gesamt_eur": _t("spalte_cashflow"),
                 "cf_kumuliert_eur": _t("spalte_kumuliert")})
    kopf, ende = _schreibe_tabelle(ws, zeile, t,
                                   {_t("spalte_cashflow"): _EUR, _t("spalte_kumuliert"): _EUR})
    balken = _balken_chart(ws, kopf, ende, [2], _t("chart_cashflow_titel"), "€")
    linie = LineChart()
    linie.add_data(Reference(ws, min_col=3, min_row=kopf, max_row=ende),
                   titles_from_data=True)
    balken += linie
    _haenge_chart(ws, balken, kopf)
    zeile = _naechste_zeile(ende, kopf)

    # --- 4. Betriebskosten (gestapelt) ---
    zeile = _sektion(ws, zeile, _t("sektion_betriebskosten"))
    opex_spalten = list(result.cashflow.opex_posten) + [
        "gemeindeabgabe_eur", "direktvermarktungskosten_eur"]
    opex_spalten = [c for c in opex_spalten if c in betrieb.columns]
    t = betrieb[["jahr"] + opex_spalten].rename(columns={
        "jahr": _t("spalte_jahr"), "gemeindeabgabe_eur": _t("spalte_gemeindeabgabe"),
        "direktvermarktungskosten_eur": _t("spalte_direktvermarktung")})
    kopf, ende = _schreibe_tabelle(
        ws, zeile, t, {c: _EUR for c in t.columns if c != _t("spalte_jahr")})
    _haenge_chart(ws, _balken_chart(
        ws, kopf, ende, list(range(2, len(t.columns) + 1)),
        _t("chart_betriebskosten_titel"), "€/Jahr", gestapelt=True), kopf)
    zeile = _naechste_zeile(ende, kopf)

    # --- 5. Finanzierung ---
    zeile = _sektion(ws, zeile, _t("sektion_kapitaldienst"))
    t = betrieb[["jahr", "zinsen_eur", "tilgung_eur", "dscr"]].rename(
        columns={"jahr": _t("spalte_jahr"), "zinsen_eur": _t("spalte_zinsen"),
                 "tilgung_eur": _t("spalte_tilgung"), "dscr": _t("spalte_dscr")})
    kopf, ende = _schreibe_tabelle(ws, zeile, t, {
        _t("spalte_zinsen"): _EUR, _t("spalte_tilgung"): _EUR, _t("spalte_dscr"): "0.00"})
    _haenge_chart(ws, _balken_chart(ws, kopf, ende, [2, 3],
                                    _t("chart_kapitaldienst_titel"), "€/Jahr", gestapelt=True),
                  kopf)
    dscr_chart = _linien_chart(ws, kopf, ende, [4], _t("chart_dscr_titel"), _t("spalte_dscr"))
    dscr_chart.y_axis.numFmt = "0.00"
    ws.add_chart(dscr_chart, f"{_CHART_SPALTE}{kopf + 18}")
    zeile = max(_naechste_zeile(ende, kopf), kopf + 37)

    # --- 6. NPV-Kurve ---
    zeile = _sektion(ws, zeile, _t("sektion_npv_kurve"))
    t = result.npv_curve.rename(columns={
        "diskontsatz_pct": _t("spalte_diskontsatz"), "npv_eur": _t("spalte_npv")})
    kopf, ende = _schreibe_tabelle(ws, zeile, t,
                                   {_t("spalte_diskontsatz"): _PCT, _t("spalte_npv"): _EUR})
    _haenge_chart(ws, _linien_chart(ws, kopf, ende, [2],
                                    _t("chart_npv_titel"), "€"), kopf)
    zeile = _naechste_zeile(ende, kopf)

    # --- 7. Sensitivitaet (Tornado) ---
    zeile = _sektion(ws, zeile, _t("sektion_sensitivitaet"))
    tornado = run_tornado(project, ga).sort_values("spanne")
    t = tornado[["name", "irr_runter", "irr_rauf"]].rename(columns={
        "name": _t("spalte_treiber"), "irr_runter": _t("spalte_irr_minus10"),
        "irr_rauf": _t("spalte_irr_plus10")})
    kopf, ende = _schreibe_tabelle(ws, zeile, t, {
        _t("spalte_irr_minus10"): _PCT, _t("spalte_irr_plus10"): _PCT})
    chart = _balken_chart(ws, kopf, ende, [2, 3],
                          _t("chart_sensitivitaet_titel"), "IRR", horizontal=True)
    chart.y_axis.numFmt = "0.0%"
    _haenge_chart(ws, chart, kopf)
    zeile = _naechste_zeile(ende, kopf)

    # --- 8. EAG-Sensitivitaet ---
    zeile = _sektion(ws, zeile, _t("sektion_eag_sensitivitaet"))
    sens = run_eag_sensitivity(project, ga)
    t = sens[["eag_zuschlagswert_ct_kwh", "equity_irr", "npv_eur"]].rename(
        columns={"eag_zuschlagswert_ct_kwh": _t("spalte_zuschlagswert_ct"),
                 "equity_irr": _t("spalte_ek_rendite"), "npv_eur": _t("spalte_npv")})
    kopf, ende = _schreibe_tabelle(ws, zeile, t, {
        _t("spalte_zuschlagswert_ct"): _CT, _t("spalte_ek_rendite"): _PCT,
        _t("spalte_npv"): _EUR})
    chart = LineChart()
    _stil(chart, _t("chart_eag_sensitivitaet_titel"), "IRR")
    chart.add_data(Reference(ws, min_col=2, min_row=kopf, max_row=ende),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=kopf + 1, max_row=ende))
    chart.y_axis.numFmt = "0.0%"
    _haenge_chart(ws, chart, kopf)
    zeile = _naechste_zeile(ende, kopf)

    # --- 9. Monte Carlo (Histogramm aus Klassen) ---
    zeile = _sektion(ws, zeile, _t("sektion_monte_carlo", n_mc=n_mc))
    mc = run_monte_carlo(project, ga, n_laeufe=n_mc,
                         sigmas=dict(MC_STANDARD_SIGMAS))
    irr = mc.irr[~np.isnan(mc.irr)]
    p10, p50, p90 = (float(np.percentile(irr, q)) for q in (10, 50, 90))
    kanten = np.linspace(irr.min(), irr.max(), 16)
    haeufig, _ = np.histogram(irr, bins=kanten)
    t = pd.DataFrame({
        _t("spalte_ek_rendite_klassenmitte"): (kanten[:-1] + kanten[1:]) / 2,
        _t("spalte_anzahl_laeufe"): haeufig,
    })
    kopf, ende = _schreibe_tabelle(ws, zeile, t, {
        _t("spalte_ek_rendite_klassenmitte"): _PCT})
    _haenge_chart(ws, _balken_chart(ws, kopf, ende, [2],
                                    _t("chart_mc_titel"),
                                    _t("spalte_anzahl_laeufe")), kopf)
    for i, (name, wert) in enumerate(
        (("P10", p10), ("Median", p50), ("P90", p90))
    ):
        ws.cell(row=ende + 2 + i, column=1, value=name).font = _F_TEXT
        z = ws.cell(row=ende + 2 + i, column=2, value=wert)
        z.font = _F_TEXT
        z.number_format = _PCT
    zeile = _naechste_zeile(ende + 5, kopf)

    # --- 10. Szenarienvergleich ---
    zeile = _sektion(ws, zeile, _t("sektion_szenarien"))
    vergleich = run_scenario_comparison(project, ga)
    kz = vergleich.kennzahlen.rename(columns={
        "szenario": _t("spalte_szenario"), "equity_irr": _t("spalte_ek_rendite"),
        "npv_eur": _t("spalte_npv"), "erloes_summe_eur": _t("spalte_erloese_gesamt")})
    kz = kz[[c for c in (_t("spalte_szenario"), _t("spalte_ek_rendite"),
                         _t("spalte_npv"), _t("spalte_erloese_gesamt"))
             if c in kz.columns]]
    kopf, ende = _schreibe_tabelle(ws, zeile, kz, {
        _t("spalte_ek_rendite"): _PCT, _t("spalte_npv"): _EUR,
        _t("spalte_erloese_gesamt"): _EUR})
    chart = _balken_chart(ws, kopf, ende, [2], _t("chart_szenarien_titel"), "IRR")
    chart.y_axis.numFmt = "0.0%"
    _haenge_chart(ws, chart, kopf)


# ---------------------------------------------------------------------------
# Uebersichtsblatt + Gesamtexport
# ---------------------------------------------------------------------------


def pipeline_ergebnis_excel(
    projekte: list[tuple[PVProject, str]],
    ga: GlobalAssumptions,
    n_mc: int = 300,
) -> bytes:
    """Erstellt die Ergebnis-Arbeitsmappe: Blatt 'Übersicht' plus je
    Projekt ein Reiter mit allen Auswertungen als native Excel-
    Diagramme. `projekte`: Liste (Projekt, gewuenschter Blattname)."""
    wb = Workbook()
    ws = wb.active
    ws.title = _t("blatt_uebersicht")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    for sp in "BCDEFGHI":
        ws.column_dimensions[sp].width = 14
    ws.cell(row=1, column=1, value=_t("titel_uebersicht")).font = _F_TITEL

    zeilen = []
    ergebnisse: list[tuple[PVProject, str]] = []
    vergeben = {_t("blatt_uebersicht")}
    for project, wunschname in projekte:
        kpis = run_valuation(project, ga).kpis
        zeilen.append({
            _t("spalte_projekt"): project.name,
            _t("spalte_typ"): (_t("typ_agri")
                               if project.anlagentyp == AnlagenTyp.AGRI_PV
                               else _t("typ_konventionell")),
            _t("spalte_aktiv"): _t("spalte_ja") if project.aktiv else _t("spalte_nein"),
            _t("spalte_kwp"): project.nennleistung_kwp,
            _t("spalte_capex"): kpis.capex_total_eur,
            _t("spalte_eigenkapital"): kpis.eigenkapital_eur,
            _t("spalte_ek_rendite"): kpis.equity_irr,
            _t("spalte_npv"): kpis.npv_eur,
            _t("spalte_dscr_min"): kpis.dscr_min,
        })
        ergebnisse.append((project, _blattname(wunschname, vergeben)))

    uebersicht = pd.DataFrame(zeilen)
    kopf, ende = _schreibe_tabelle(ws, 3, uebersicht, {
        _t("spalte_kwp"): "#,##0", _t("spalte_capex"): _EUR,
        _t("spalte_eigenkapital"): _EUR, _t("spalte_ek_rendite"): _PCT,
        _t("spalte_npv"): _EUR, _t("spalte_dscr_min"): "0.00"})
    chart = BarChart()
    chart.type = "col"
    _stil(chart, _t("chart_irr_je_projekt_titel"), "IRR")
    chart.add_data(Reference(ws, min_col=7, min_row=kopf, max_row=ende),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=kopf + 1,
                                   max_row=ende))
    chart.y_axis.numFmt = "0.0%"
    ws.add_chart(chart, f"A{ende + 3}")
    ws.cell(row=ende + 21, column=1,
            value=_t("hinweis_uebersicht")).font = _F_TEXT

    for project, blatt in ergebnisse:
        _projektblatt(wb.create_sheet(blatt), project, ga, n_mc)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
